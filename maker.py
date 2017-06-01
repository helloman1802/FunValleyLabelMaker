import openpyxl
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import PyPDF2
import os
import pyqrcode as qr
from reportlab.pdfgen import canvas
from random_words import RandomWords
from random import randint
rw = RandomWords()

# Google Sheets loging using O Auth 2.0
scope = ['https://spreadsheets.google.com/feeds']
credentials = ServiceAccountCredentials.from_json_keyfile_name('FunValleyLabelMaker.json', scope)
gc = gspread.authorize(credentials)
sh = gc.open("FunValleyGiftshopItemData")
worksheet = sh.worksheet("Sheet")






def user_input():
    fun_sku_yes_no = input("Fun Valley sku; Yes or No: ")

    if fun_sku_yes_no.upper() == "YES" or fun_sku_yes_no.upper() == "Y":
        fun_sku_number = input("What is the sku?: ")
        fun_sku_number = fun_sku_number.upper()
        if fun_sku_number[:2] == "FV":
            row_values = []
            c = worksheet.find(fun_sku_number)
            row, col = c.row, c.col
            for i in range(9):
                row_values.append(worksheet.cell(row, i+1).value)

            print(row_values)
            print_labels = input("Print labels; Yes or No: ")
            print_labels = print_labels.upper()
            if print_labels == "YES" or print_labels == "Y":
                circle_or_square = worksheet.cell(row, 10).value
                if int(circle_or_square) == 0:
                    print("Printing on 3x4 circle labels")
                    logo = input("Print logo overlay: Yes or No: ")

                    if logo.upper() == "YES" or logo.upper() == "Y":
                        overlay = "overlay.pdf"
                    elif logo.upper() == "NO" or logo.upper() == "N":
                        overlay = "overlay3.pdf"
                    else:
                        print("GENERAL FAILURE")
                    position = input("Pick position 1-12: ")
                    quantity = input("Choose quantity: ")
                    counter = int(position)
                    quantity = int(quantity)


                    c = canvas.Canvas('watermark.pdf')

                    n = 0
                    t = 1

                    x_qr = 101
                    y_qr = 626

                    x_price = 116
                    y_price = 695

                    x_dis = 116
                    y_dis = 656

                    page_num_x = 20
                    page_num_y = 20
                    page_counter = 0
                    word = rw.random_word()


                    price =       worksheet.cell(row, 5).value
                    discription = worksheet.cell(row, 2).value
                    size  =       worksheet.cell(row, 4).value
                    color =       worksheet.cell(row, 3).value
                    while True:
                        file_name = "qrcode%d.png" % (n)
                        QR = qr.create(fun_sku_number)
                        QR.png(file_name, scale=1)




                        if counter == 1:
                            #c.drawImage("welcome_925.png", -11 , 630)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis, y_dis, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price, y_price, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_dis,y_dis+24,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_dis,y_dis+12,color)
                            c.drawImage(file_name, x_qr, y_qr)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break

                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 2:
                            #c.drawImage("welcome_925.png", 250, 740)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 190, y_dis, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 190, y_price, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_dis + 190,y_dis+24,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_dis + 190,y_dis+12,color)
                            c.drawImage(file_name, x_qr + 190, y_qr)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 3:
                            #c.drawImage("welcome_925.png", x_price + 400, 740)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 380, y_dis, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 380, y_price, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 380,y_dis+24,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 380,y_dis+12,color)
                            c.drawImage(file_name, x_qr + 380, y_qr)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 4:
                            #c.drawImage("welcome_925.png", 37.5, 540)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis, y_dis - 183, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price, y_price - 185, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_dis, y_dis-160,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_dis,y_dis-172,color)
                            c.drawImage(file_name, x_qr, y_qr - 185)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 5:
                            #c.drawImage("welcome_925.png", 250, 540)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 190, y_dis - 183, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 190, y_price - 185, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 190, y_dis-160,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 190,y_dis-172,color)
                            c.drawImage(file_name, x_qr + 190, y_qr - 185)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 6:
                            #c.drawImage("welcome_925.png", x_price + 400, 540)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 380, y_dis - 183, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 380, y_price - 185, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 380, y_dis-160,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 380,y_dis-172,color)
                            c.drawImage(file_name, x_qr+380, y_qr - 185)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 7:
                            #c.drawImage("welcome_925.png", 37.5, 340)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis, y_dis - 368, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price, y_price - 370, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_dis, y_dis - 346,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_dis,y_dis-357,color)
                            c.drawImage(file_name, x_qr, y_qr - 370)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 8:
                            #c.drawImage("welcome_925.png", 250, 340)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis+190, y_dis - 368, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price+190, y_price - 370, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price+190, y_dis - 346,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price+190,y_dis-357,color)
                            c.drawImage(file_name, x_qr + 190, y_qr - 370)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 9:
                            #c.drawImage("welcome_925.png", x_price + 400, 340)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 380, y_dis - 368, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 380, y_price - 370, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 380,  y_dis - 346,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 380,y_dis-357,color)
                            c.drawImage(file_name, x_qr+380, y_qr - 370)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 10:
                            #c.drawImage("welcome_925.png", 37.5, 140)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis, y_dis - 553, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price, y_price-555, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_dis, y_dis-530,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_dis,y_dis-542,color)
                            c.drawImage(file_name, x_qr, y_qr-555)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 11:
                            #c.drawImage("welcome_925.png", 250, 140)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 190, y_dis - 553, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 190, y_price-555, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 190,y_dis-530,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 190,y_dis-542,color)
                            c.drawImage(file_name, x_qr + 190, y_qr-555)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open(overlay, 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        elif counter == 12:
                            #c.drawImage("welcome_925.png", x_price + 400, 140)
                            c.setFont('Helvetica', 8)
                            c.drawCentredString(x_dis + 380, y_dis - 553, discription)
                            c.setFont('Helvetica-Bold', 15)
                            c.drawCentredString(x_price + 380, y_price-555, price)
                            c.setFont('Helvetica-Bold', 13)
                            if size:
                                c.drawCentredString(x_price + 380, y_dis-530,size)
                            c.setFont('Helvetica-Bold', 10)
                            if color:
                                c.drawCentredString(x_price + 380,y_dis-542,color)
                            c.drawImage(file_name, x_qr+380, y_qr-555)
                            counter = 1
                            page_counter = page_counter + 1
                            page_number = "%s %d" % (word, page_counter)
                            c.drawString(page_num_x, page_num_y, page_number)
                            c.save()
                            overlayFile = open(overlay, 'rb')
                            pdfReader = PyPDF2.PdfFileReader(overlayFile)
                            minutesFirstPage = pdfReader.getPage(0)
                            pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                            minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                            pdfWriter = PyPDF2.PdfFileWriter()
                            pdfWriter.addPage(minutesFirstPage)

                            for pageNum in range(1, pdfReader.numPages):
                                    pageObj = pdfReader.getPage(pageNum)
                                    pdfWriter.addPage(pageObj)
                            resultPdfFile = open('merged.pdf', 'wb')
                            pdfWriter.write(resultPdfFile)
                            overlayFile.close()
                            resultPdfFile.close()
                            PDF_name = 'watermark.pdf'
                            c = canvas.Canvas(PDF_name)
                            os.system("lp merged.pdf")
                            input("Waiting")


                            t = t+1
                            print("t: ", t, "Quantity: ", quantity)

                        if n > 20:
                            n = 0

                if int(circle_or_square) == 1:
                    print("Printing on 4x20 square labels")
                    position = input("Pick position 1-80: ")
                    quantity = input("Choose quantity: ")
                    counter = int(position)
                    counter = counter - 1
                    quantity = int(quantity)

                    c = canvas.Canvas('watermark.pdf')

                    n = 0
                    t = 1

                    x = 170
                    y = 36.4

                    discr_x = 40
                    discr_y = 36.4

                    color_x = 40
                    color_y = 36.4

                    price_x = 40
                    price_y = 36.4

                    page_num_x = 20
                    page_num_y = 20
                    page_counter = 0
                    word = rw.random_word()


                    price =       worksheet.cell(row, 5).value
                    discription = worksheet.cell(row, 2).value
                    size  =       worksheet.cell(row, 4).value
                    color =       worksheet.cell(row, 3).value

                    while True:
                        file_name = "qrcode%d.png" % (n)
                        QR = qr.create(fun_sku_number)
                        QR.png(file_name, scale=1)

                        if counter == 0:
                            c.drawImage(file_name, 25, 37.5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, 38.5, discription)
                            if color:
                                c.drawString(color_x+15, color_y+9.4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, price_y+20, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 1:
                            c.drawImage(file_name, x+5, 37.5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, 38.5, discription)
                            if color:
                                c.drawString(color_x+165, color_y+9.4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, (price_y+20), price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 2:
                            c.drawImage(file_name, x+155, 37.5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, 38.5, discription)
                            if color:
                                c.drawString(color_x+320, color_y+9.4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, (price_y+20), price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 3:
                            c.drawImage(file_name, x+300, 37.5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, 38.5, discription)
                            if color:
                                c.drawString(color_x+465, color_y+9.4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, (price_y+20), price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 4:
                            c.drawImage(file_name, 25, y*2)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*2, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*2, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*2, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 5:
                            c.drawImage(file_name, x+5, y*2)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*2, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*2, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*2, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 6:
                            c.drawImage(file_name, x+155, y*2)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*2, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*2, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*2, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 7:
                            c.drawImage(file_name, x+300, y*2)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*2, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*2, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*2, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 8:
                            c.drawImage(file_name, 25, y*3)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*3, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*3, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*3, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 9:
                            c.drawImage(file_name, x+5, y*3)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*3, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*3, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*3, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 10:
                            c.drawImage(file_name, x+155, y*3)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*3, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*3, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*3, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 11:
                            c.drawImage(file_name, x+300, y*3)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*3, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*3, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*3, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 12:
                            c.drawImage(file_name, 25, y*4)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*4, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*4, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 13:
                            c.drawImage(file_name, x+5, y*4)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*4, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*4, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 14:
                            c.drawImage(file_name, x+155, y*4)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*4, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*4, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 15:
                            c.drawImage(file_name, x+300, y*4)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*4, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*4, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*4, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1


                        elif counter == 16:
                            c.drawImage(file_name, 25, y*5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*5, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*5, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*5, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 17:
                            c.drawImage(file_name, x+5, y*5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*5, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*5, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*5, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 18:
                            c.drawImage(file_name, x+155, y*5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*5, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*5, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*5, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 19:
                            c.drawImage(file_name, x+300, y*5)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*5, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*5, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*5, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 20:
                            c.drawImage(file_name, 25, y*6)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*6, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*6, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*6, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 21:
                            c.drawImage(file_name, x+5, y*6)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*6, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*6, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*6, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 22:
                            c.drawImage(file_name, x+155, y*6)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*6, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*6, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*6, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 23:
                            c.drawImage(file_name, x+300, y*6)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*6, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*6, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*6, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 24:
                            c.drawImage(file_name, 25, y*7)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*7, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*7, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*7, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 25:
                            c.drawImage(file_name, x+5, y*7)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*7, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*7, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*7, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 26:
                            c.drawImage(file_name, x+155, y*7)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*7, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*7, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*7, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 27:
                            c.drawImage(file_name, x+300, y*7)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*7, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*7, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*7, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 28:
                            c.drawImage(file_name, 25, y*8)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*8, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*8, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*8, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 29:
                            c.drawImage(file_name, x+5, y*8)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*8, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*8, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*8, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 30:
                            c.drawImage(file_name, x+155, y*8)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*8, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*8, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*8, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 31:
                            c.drawImage(file_name, x+300, y*8)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*8, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*8, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*8, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 32:
                            c.drawImage(file_name, 25, y*9)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*9, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*9, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*9, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 33:
                            c.drawImage(file_name, x+5, y*9)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*9, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*9, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*9, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 34:
                            c.drawImage(file_name, x+155, y*9)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*9, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*9, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*9, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 35:
                            c.drawImage(file_name, x+300, y*9)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*9, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*9, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*9, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 36:
                            c.drawImage(file_name, 25, y*10)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*10, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*10, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*10, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 37:
                            c.drawImage(file_name, x+5, y*10)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*10, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*10, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*10, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 38:
                            c.drawImage(file_name, x+155, y*10)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*10, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*10, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*10, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 39:
                            c.drawImage(file_name, x+300, y*10)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*10, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*10, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*10, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 40:
                            c.drawImage(file_name, 25, y*11)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*11, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*11, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*11, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 41:
                            c.drawImage(file_name, x+5, y*11)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*11, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*11, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*11, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 42:
                            c.drawImage(file_name, x+155, y*11)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*11, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*11, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*11, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 43:
                            c.drawImage(file_name, x+300, y*11)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*11, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*11, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*11, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 44:
                            c.drawImage(file_name, 25, y*12)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*12, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*12, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*12, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 45:
                            c.drawImage(file_name, x+5, y*12)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*12, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*12, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*12, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 46:
                            c.drawImage(file_name, x+155, y*12)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*12, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*12, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*12, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 47:
                            c.drawImage(file_name, x+300, y*12)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*12, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*12, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*12, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 48:
                            c.drawImage(file_name, 25, y*13)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*13, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*13, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*13, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 49:
                            c.drawImage(file_name, x+5, y*13)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*13, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*13, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*13, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 50:
                            c.drawImage(file_name, x+155, y*13)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*13, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*13, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*13, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 51:
                            c.drawImage(file_name, x+300, y*13)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*13, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*13, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*13, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 52:
                            c.drawImage(file_name, 25, y*14)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*14, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*14, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*14, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 53:
                            c.drawImage(file_name, x+5, y*14)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*14, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*14, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*14, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 54:
                            c.drawImage(file_name, x+155, y*14)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*14, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*14, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*14, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 55:
                            c.drawImage(file_name, x+300, y*14)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*14, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*14, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*14, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 56:
                            c.drawImage(file_name, 25, y*15)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*15, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*15, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*15, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 57:
                            c.drawImage(file_name, x+5, y*15)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*15, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*15, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*15, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 58:
                            c.drawImage(file_name, x+155, y*15)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*15, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*15, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*15, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 59:
                            c.drawImage(file_name, x+300, y*15)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*15, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*15, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*15, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 60:
                            c.drawImage(file_name, 25, y*16)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*16, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*16, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*16, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 61:
                            c.drawImage(file_name, x+5, y*16)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*16, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*16, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*16, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 62:
                            c.drawImage(file_name, x+155, y*16)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*16, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*16, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*16, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 63:
                            c.drawImage(file_name, x+300, y*16)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*16, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*16, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*16, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 64:
                            c.drawImage(file_name, 25, y*17)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*17, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*17, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*17, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 65:
                            c.drawImage(file_name, x+5, y*17)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*17, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*17, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*17, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 66:
                            c.drawImage(file_name, x+155, y*17)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*17, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*17, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*17, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 67:
                            c.drawImage(file_name, x+300, y*17)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*17, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*17, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*17, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 68:
                            c.drawImage(file_name, 25, y*18)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*18, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*18, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*18, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 69:
                            c.drawImage(file_name, x+5, y*18)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*18, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*18, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*18, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 70:
                            c.drawImage(file_name, x+155, y*18)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*18, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*18, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*18, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 71:
                            c.drawImage(file_name, x+300, y*18)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*18, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*18, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*18, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 72:
                            c.drawImage(file_name, 25, y*19)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*19, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*19, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*19, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 73:
                            c.drawImage(file_name, x+5, y*19)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*19, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*19, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*19, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 74:
                            c.drawImage(file_name, x+155, y*19)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*19, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*19, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*19, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 75:
                            c.drawImage(file_name, x+300, y*19)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*19, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*19, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*19, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1

                        elif counter == 76:
                            c.drawImage(file_name, 25, y*20)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+15, discr_y*20, discription)
                            if color:
                                c.drawString(color_x+15, 7+color_y*20, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+15, 18+price_y*20, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 77:
                            c.drawImage(file_name, x+5, y*20)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+165, discr_y*20, discription)
                            if color:
                                c.drawString(color_x+165, 7+color_y*20, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+165, 18+price_y*20, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 78:
                            c.drawImage(file_name, x+155, y*20)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+320, discr_y*20, discription)
                            if color:
                                c.drawString(color_x+320, 7+color_y*20, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+320, 18+price_y*20, price)
                            if t == quantity:
                                page_counter = page_counter + 1
                                page_number = "%s %d" % (word, page_counter)
                                c.drawString(page_num_x, page_num_y, page_number)
                                c.save()
                                overlayFile = open('overlay2.pdf', 'rb')
                                pdfReader = PyPDF2.PdfFileReader(overlayFile)
                                minutesFirstPage = pdfReader.getPage(0)
                                pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                                minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                                pdfWriter = PyPDF2.PdfFileWriter()
                                pdfWriter.addPage(minutesFirstPage)

                                for pageNum in range(1, pdfReader.numPages):
                                        pageObj = pdfReader.getPage(pageNum)
                                        pdfWriter.addPage(pageObj)
                                resultPdfFile = open('merged.pdf', 'wb')
                                pdfWriter.write(resultPdfFile)
                                overlayFile.close()
                                resultPdfFile.close()
                                os.system("lp merged.pdf")
                                break
                            counter = counter +1

                            t = t+1
                        elif counter == 79:
                            c.drawImage(file_name, x+300, y*20)
                            c.setFont('Helvetica', 6)
                            c.drawString(discr_x+465, discr_y*20, discription)
                            if color:
                                c.drawString(color_x+465, 7+color_y*20, color)
                            c.setFont('Helvetica-Bold', 10)
                            c.drawString(price_x+465, 18+price_y*20, price)
                            page_counter = page_counter + 1
                            page_number = "%s %d" % (word, page_counter)
                            c.drawString(page_num_x, page_num_y, page_number)
                            counter = 0

                            t = t+1
                            c.save()
                            overlayFile = open('overlay2.pdf', 'rb')
                            pdfReader = PyPDF2.PdfFileReader(overlayFile)
                            minutesFirstPage = pdfReader.getPage(0)
                            pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf', 'rb'))
                            minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))
                            pdfWriter = PyPDF2.PdfFileWriter()
                            pdfWriter.addPage(minutesFirstPage)

                            for pageNum in range(1, pdfReader.numPages):
                                    pageObj = pdfReader.getPage(pageNum)
                                    pdfWriter.addPage(pageObj)
                            resultPdfFile = open('merged.pdf', 'wb')
                            pdfWriter.write(resultPdfFile)
                            overlayFile.close()
                            resultPdfFile.close()
                            PDF_name = 'watermark.pdf'
                            c = canvas.Canvas(PDF_name)


                            os.system("lp merged.pdf")

                            input("Waiting")


                        if n > 100:
                            n = 0






            else:
                print("GENERAL FAILURE 1")
        else:
            print("GENERAL FAILURE 2")

    elif fun_sku_yes_no.upper() == "NO" or fun_sku_yes_no.upper() == "N":
        print("Creating New Sku")
        random_number = randint(0,9999)
        fun_sku_number = "FV%04d" % (random_number,)
        while True:
            try:
                raise Exception(worksheet.find(fun_sku_number))
            except Exception:
                print("Unique SKU")
                worksheet.insert_row([fun_sku_number, "Pickled Pickles"], index=913)
                break


user_input()
